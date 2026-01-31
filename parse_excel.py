"""
Excel Parser Script
This script parses the excel_app.xlsx file and extracts all details including:
- Sheet names
- Tables and their structure
- Cell values and formulas
- Cell dependencies
- Conditional formatting
- Data validation
- Charts/plots
- Named ranges
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from pathlib import Path


def parse_excel(filepath: str) -> dict:
    """
    Parse an Excel file and extract comprehensive information.
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        Dictionary containing all parsed data
    """
    # Load workbook with data_only=False to get formulas
    wb = load_workbook(filepath, data_only=False)
    
    # Load another instance with data_only=True to get calculated values
    wb_values = load_workbook(filepath, data_only=True)
    
    result = {
        "filename": Path(filepath).name,
        "sheets": [],
        "named_ranges": [],
        "defined_names": []
    }
    
    # Extract defined names (named ranges)
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        result["defined_names"].append({
            "name": name,
            "value": dn.value if hasattr(dn, 'value') else str(dn),
            "scope": dn.localSheetId if hasattr(dn, 'localSheetId') else None
        })
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_values = wb_values[sheet_name]
        
        sheet_data = {
            "name": sheet_name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_cells": [str(mc) for mc in ws.merged_cells.ranges],
            "tables": [],
            "cells": [],
            "conditional_formatting": [],
            "data_validation": [],
            "charts": [],
            "row_dimensions": {},
            "col_dimensions": {}
        }
        
        # Extract tables
        for table_name, table in ws.tables.items():
            sheet_data["tables"].append({
                "name": table_name,
                "ref": table.ref,
                "display_name": table.displayName
            })
        
        # Extract cell data (only cells with values or formulas)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None or cell.data_type != 'n':
                    cell_info = {
                        "address": cell.coordinate,
                        "row": cell.row,
                        "col": cell.column,
                        "col_letter": get_column_letter(cell.column),
                        "value": None,
                        "formula": None,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "fill_color": None,
                        "font_color": None,
                        "alignment": None,
                        "border": None
                    }
                    
                    # Get the value
                    if cell.data_type == 'f':
                        cell_info["formula"] = str(cell.value)
                        # Get calculated value from the values-only workbook
                        value_cell = ws_values[cell.coordinate]
                        cell_info["value"] = value_cell.value
                    else:
                        cell_info["value"] = cell.value
                    
                    # Get fill color
                    if cell.fill and cell.fill.fgColor:
                        if cell.fill.fgColor.type == 'rgb' and cell.fill.fgColor.rgb:
                            cell_info["fill_color"] = cell.fill.fgColor.rgb
                        elif cell.fill.fgColor.type == 'theme':
                            cell_info["fill_color"] = f"theme:{cell.fill.fgColor.theme}"
                    
                    # Get font color
                    if cell.font and cell.font.color:
                        if cell.font.color.type == 'rgb' and cell.font.color.rgb:
                            cell_info["font_color"] = cell.font.color.rgb
                        elif cell.font.color.type == 'theme':
                            cell_info["font_color"] = f"theme:{cell.font.color.theme}"
                    
                    # Get alignment
                    if cell.alignment:
                        cell_info["alignment"] = {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text
                        }
                    
                    sheet_data["cells"].append(cell_info)
        
        # Extract conditional formatting
        for cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
            for rule in cf_rules:
                cf_data = {
                    "range": str(cf_range),
                    "type": rule.type,
                    "priority": rule.priority
                }
                if hasattr(rule, 'formula') and rule.formula:
                    cf_data["formula"] = list(rule.formula)
                if rule.dxf:
                    if rule.dxf.fill:
                        cf_data["fill"] = str(rule.dxf.fill.fgColor.rgb) if rule.dxf.fill.fgColor else None
                    if rule.dxf.font:
                        cf_data["font_color"] = str(rule.dxf.font.color.rgb) if rule.dxf.font and rule.dxf.font.color else None
                sheet_data["conditional_formatting"].append(cf_data)
        
        # Extract data validation
        for dv in ws.data_validations.dataValidation:
            dv_data = {
                "ranges": str(dv.sqref),
                "type": dv.type,
                "allow_blank": dv.allow_blank,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "operator": dv.operator,
                "prompt_title": dv.promptTitle,
                "prompt": dv.prompt,
                "error_title": dv.errorTitle,
                "error": dv.error
            }
            sheet_data["data_validation"].append(dv_data)
        
        # Extract charts
        for chart in ws._charts:
            chart_data = {
                "type": type(chart).__name__,
                "title": str(chart.title) if chart.title else None,
                "anchor": str(chart.anchor) if hasattr(chart, 'anchor') else None
            }
            # Try to get series data
            if hasattr(chart, 'series'):
                chart_data["series"] = []
                for series in chart.series:
                    series_info = {}
                    if hasattr(series, 'val') and series.val:
                        series_info["values"] = series.val.numRef.f if series.val.numRef else None
                    if hasattr(series, 'cat') and series.cat:
                        series_info["categories"] = series.cat.numRef.f if series.cat.numRef else (
                            series.cat.strRef.f if series.cat.strRef else None
                        )
                    if hasattr(series, 'title') and series.title:
                        series_info["title"] = series.title
                    chart_data["series"].append(series_info)
            sheet_data["charts"].append(chart_data)
        
        # Row dimensions (hidden rows, height)
        for row_num, row_dim in ws.row_dimensions.items():
            if row_dim.hidden or row_dim.height:
                sheet_data["row_dimensions"][row_num] = {
                    "hidden": row_dim.hidden,
                    "height": row_dim.height
                }
        
        # Column dimensions (hidden columns, width)
        for col, col_dim in ws.column_dimensions.items():
            if col_dim.hidden or col_dim.width:
                sheet_data["col_dimensions"][col] = {
                    "hidden": col_dim.hidden,
                    "width": col_dim.width
                }
        
        result["sheets"].append(sheet_data)
    
    wb.close()
    wb_values.close()
    
    return result


def analyze_dependencies(parsed_data: dict) -> dict:
    """
    Analyze cell dependencies from formulas.
    
    Args:
        parsed_data: The parsed Excel data
        
    Returns:
        Dictionary mapping cells to their dependencies
    """
    import re
    
    # Pattern to match cell references like A1, $A$1, Sheet1!A1, etc.
    cell_pattern = re.compile(r"(?:'?([^'!]+)'?!)?\$?([A-Z]+)\$?(\d+)")
    range_pattern = re.compile(r"(?:'?([^'!]+)'?!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")
    
    dependencies = {}
    
    for sheet in parsed_data["sheets"]:
        sheet_name = sheet["name"]
        
        for cell in sheet["cells"]:
            if cell["formula"]:
                formula = cell["formula"]
                cell_deps = []
                
                # Find range references first (to avoid double-matching)
                for match in range_pattern.finditer(formula):
                    ref_sheet = match.group(1) or sheet_name
                    start_col = match.group(2)
                    start_row = match.group(3)
                    end_col = match.group(4)
                    end_row = match.group(5)
                    cell_deps.append({
                        "type": "range",
                        "sheet": ref_sheet,
                        "start": f"{start_col}{start_row}",
                        "end": f"{end_col}{end_row}"
                    })
                
                # Remove ranges from formula to find individual cells
                formula_no_ranges = range_pattern.sub("", formula)
                
                # Find individual cell references
                for match in cell_pattern.finditer(formula_no_ranges):
                    ref_sheet = match.group(1) or sheet_name
                    col = match.group(2)
                    row = match.group(3)
                    cell_deps.append({
                        "type": "cell",
                        "sheet": ref_sheet,
                        "address": f"{col}{row}"
                    })
                
                if cell_deps:
                    key = f"{sheet_name}!{cell['address']}"
                    dependencies[key] = cell_deps
    
    return dependencies


def format_for_markdown(parsed_data: dict, dependencies: dict) -> str:
    """
    Format the parsed data as markdown for the documentation.
    
    Args:
        parsed_data: The parsed Excel data
        dependencies: Cell dependency data
        
    Returns:
        Formatted markdown string
    """
    md = []
    md.append(f"**File:** `{parsed_data['filename']}`\n")
    
    # Sheet overview
    md.append("## Sheet Overview\n")
    for i, sheet in enumerate(parsed_data['sheets'], 1):
        md.append(f"{i}. **{sheet['name']}** - Dimensions: {sheet['dimensions']} ({sheet['max_row']} rows x {sheet['max_col']} columns)")
    md.append("")
    
    # Defined names
    if parsed_data['defined_names']:
        md.append("## Named Ranges / Defined Names\n")
        md.append("| Name | Value | Scope |")
        md.append("|------|-------|-------|")
        for dn in parsed_data['defined_names']:
            scope = f"Sheet {dn['scope']}" if dn['scope'] is not None else "Workbook"
            md.append(f"| {dn['name']} | `{dn['value']}` | {scope} |")
        md.append("")
    
    # Detailed sheet information
    for sheet in parsed_data['sheets']:
        md.append(f"---\n\n## Sheet: {sheet['name']}\n")
        
        # Merged cells
        if sheet['merged_cells']:
            md.append(f"### Merged Cells\n")
            md.append(", ".join([f"`{mc}`" for mc in sheet['merged_cells']]))
            md.append("")
        
        # Tables
        if sheet['tables']:
            md.append("### Excel Tables\n")
            md.append("| Table Name | Range | Display Name |")
            md.append("|------------|-------|--------------|")
            for table in sheet['tables']:
                md.append(f"| {table['name']} | {table['ref']} | {table['display_name']} |")
            md.append("")
        
        # Charts
        if sheet['charts']:
            md.append("### Charts\n")
            for chart in sheet['charts']:
                md.append(f"- **{chart['type']}**")
                if chart['title']:
                    md.append(f"  - Title: {chart['title']}")
                if 'series' in chart and chart['series']:
                    md.append(f"  - Series:")
                    for series in chart['series']:
                        if series.get('title'):
                            md.append(f"    - {series.get('title')}")
                        if series.get('values'):
                            md.append(f"      - Values: `{series['values']}`")
                        if series.get('categories'):
                            md.append(f"      - Categories: `{series['categories']}`")
            md.append("")
        
        # Data validation
        if sheet['data_validation']:
            md.append("### Data Validation Rules\n")
            md.append("| Ranges | Type | Formula/Options |")
            md.append("|--------|------|-----------------|")
            for dv in sheet['data_validation']:
                formula = dv['formula1'] or ""
                md.append(f"| {dv['ranges']} | {dv['type']} | `{formula}` |")
            md.append("")
        
        # Conditional formatting
        if sheet['conditional_formatting']:
            md.append("### Conditional Formatting\n")
            md.append("| Range | Type | Formula | Fill Color |")
            md.append("|-------|------|---------|------------|")
            for cf in sheet['conditional_formatting']:
                formula = cf.get('formula', [''])[0] if cf.get('formula') else ''
                fill = cf.get('fill', '-')
                md.append(f"| {cf['range']} | {cf['type']} | `{formula}` | {fill} |")
            md.append("")
        
        # Cells grouped by area (detecting table-like regions)
        md.append("### Cell Data\n")
        
        # Separate cells with formulas from static values
        formula_cells = [c for c in sheet['cells'] if c['formula']]
        value_cells = [c for c in sheet['cells'] if c['value'] is not None and not c['formula']]
        
        # Group cells by row ranges to identify tables
        if value_cells:
            md.append("#### Static Values (Sample - User Input Areas)\n")
            md.append("| Address | Value | Fill Color | Data Type |")
            md.append("|---------|-------|------------|-----------|")
            
            # Get colored cells (likely input cells)
            colored_cells = [c for c in value_cells if c['fill_color'] and c['fill_color'] not in ['00000000', 'FFFFFFFF', '00000000']]
            
            # Show all colored cells
            for cell in colored_cells:
                value = str(cell['value'])[:50] + "..." if len(str(cell['value'])) > 50 else cell['value']
                md.append(f"| {cell['address']} | {value} | {cell['fill_color']} | {cell['data_type']} |")
            md.append("")
        
        if formula_cells:
            md.append("#### Formulas\n")
            md.append("| Address | Formula | Calculated Value |")
            md.append("|---------|---------|------------------|")
            
            # Show all formulas
            for cell in formula_cells:
                formula = cell['formula']
                value = str(cell['value']) if cell['value'] is not None else "-"
                md.append(f"| {cell['address']} | `{formula}` | {value} |")
            md.append("")
    
    # Cell dependencies summary
    if dependencies:
        md.append("---\n\n## Cell Dependencies\n")
        md.append("This section shows which cells depend on other cells through formulas.\n")
        
        # Group by sheet
        deps_by_sheet = {}
        for cell_key, deps in dependencies.items():
            sheet_name = cell_key.split("!")[0]
            if sheet_name not in deps_by_sheet:
                deps_by_sheet[sheet_name] = {}
            deps_by_sheet[sheet_name][cell_key] = deps
        
        for sheet_name, sheet_deps in deps_by_sheet.items():
            md.append(f"### {sheet_name}\n")
            
            # Show all dependencies
            for cell_key, deps in sheet_deps.items():
                dep_strs = []
                for dep in deps:
                    if dep["type"] == "cell":
                        if dep["sheet"] != sheet_name:
                            dep_strs.append(f"{dep['sheet']}!{dep['address']}")
                        else:
                            dep_strs.append(dep['address'])
                    else:
                        if dep["sheet"] != sheet_name:
                            dep_strs.append(f"{dep['sheet']}!{dep['start']}:{dep['end']}")
                        else:
                            dep_strs.append(f"{dep['start']}:{dep['end']}")
                
                cell_addr = cell_key.split("!")[1]
                md.append(f"- `{cell_addr}` ← {', '.join([f'`{d}`' for d in dep_strs])}")
            md.append("")
    
    return "\n".join(md)


def main():
    """Main function to parse Excel and generate documentation."""
    excel_path = Path(__file__).parent / "legacy" / "excel_app.xlsx"
    output_path = Path(__file__).parent / "docs" / "the_excel_i_have.md"
    
    print(f"Parsing Excel file: {excel_path}")
    
    # Parse the Excel file
    parsed_data = parse_excel(str(excel_path))
    
    # Analyze dependencies
    dependencies = analyze_dependencies(parsed_data)
    
    # Format as markdown
    markdown_content = format_for_markdown(parsed_data, dependencies)
    
    # Read the existing file and find the ### Data section
    with open(output_path, 'r', encoding='utf-8') as f:
        existing_content = f.read()
    
    # Find the ### Data section and replace everything after it
    data_marker = "### Data"
    if data_marker in existing_content:
        # Split at the marker and keep everything before it (including the marker)
        parts = existing_content.split(data_marker)
        new_content = parts[0] + data_marker + "\n\n" + markdown_content
    else:
        # Append to the end if marker not found
        new_content = existing_content + "\n\n" + data_marker + "\n\n" + markdown_content
    
    # Write the updated content
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(new_content)
    
    print(f"Documentation updated: {output_path}")
    
    # Also save the raw JSON for reference
    json_path = Path(__file__).parent / "docs" / "excel_parsed_data.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        # Convert values to strings for JSON serialization
        def serialize(obj):
            if hasattr(obj, '__dict__'):
                return str(obj)
            return obj
        json.dump(parsed_data, f, indent=2, default=str)
    
    print(f"Raw data saved: {json_path}")


if __name__ == "__main__":
    main()
